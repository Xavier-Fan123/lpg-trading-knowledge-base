import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime, timedelta
import os

wb = openpyxl.Workbook()

# ── Style definitions ──
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NORMAL_FONT = Font(name="Calibri", size=10)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_range(ws, start_row, end_row, max_col, fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════
# 1. CARGO SCHEDULE + TANK INVENTORY
# ════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1-Cargo Schedule"
ws1.sheet_properties.tabColor = "2F5496"

headers1 = [
    "Cargo #", "Month", "Product", "Volume (MT)", "Loading Window",
    "Vessel Name", "Vessel cbm", "ETA Load", "Buyer", "Destination Port",
    "Pricing Formula", "Pricing Month", "Layday", "Cancelling",
    "Status", "B/L Date", "Invoice Date", "Payment Due", "Remarks"
]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers1))
set_col_widths(ws1, [9, 9, 10, 12, 14, 14, 10, 12, 16, 14, 16, 12, 12, 12, 12, 12, 12, 12, 20])

statuses = ["Planned", "Nominated", "Loading", "Sailing", "Discharged", "Invoiced", "Settled"]
from openpyxl.worksheet.datavalidation import DataValidation
dv_status = DataValidation(type="list", formula1='"' + ",".join(statuses) + '"', allow_blank=True)
dv_status.error = "Pick a valid status"
ws1.add_data_validation(dv_status)
dv_status.add("O2:O50")

dv_product = DataValidation(type="list", formula1='"Propane,Butane,Mix"', allow_blank=True)
ws1.add_data_validation(dv_product)
dv_product.add("C2:C50")

style_range(ws1, 2, 25, len(headers1), INPUT_FILL)

# Tank inventory sub-section (row 28+)
ws1.cell(row=28, column=1, value="TANK INVENTORY TRACKER").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
tank_headers = ["Date", "Opening (MT)", "Production (MT)", "Liftings (MT)", "Closing (MT)",
                "Tank Capacity (MT)", "Utilization %", "Days to Tank-Top", "Alert"]
for c, h in enumerate(tank_headers, 1):
    ws1.cell(row=29, column=c, value=h)
style_header(ws1, 29, len(tank_headers))

for r in range(30, 60):
    ws1.cell(row=r, column=5).value = f"=B{r}+C{r}-D{r}"  # Closing = Open + Prod - Lift
    ws1.cell(row=r, column=5).fill = CALC_FILL
    ws1.cell(row=r, column=7).value = f'=IF(F{r}>0,E{r}/F{r},"")'  # Utilization
    ws1.cell(row=r, column=7).fill = CALC_FILL
    ws1.cell(row=r, column=7).number_format = '0.0%'
    ws1.cell(row=r, column=8).value = f'=IF(C{r}>0,(F{r}-E{r})/C{r},"")'  # Days to tank-top
    ws1.cell(row=r, column=8).fill = CALC_FILL
    ws1.cell(row=r, column=9).value = f'=IF(AND(G{r}<>"",G{r}>0.85),"⚠ HIGH","OK")'
    ws1.cell(row=r, column=9).fill = CALC_FILL

style_range(ws1, 30, 59, len(tank_headers), INPUT_FILL)
for r in range(30, 60):
    for c in [5, 7, 8, 9]:
        ws1.cell(row=r, column=c).fill = CALC_FILL

ws1.freeze_panes = "A2"


# ════════════════════════════════════════════
# 2. POSITION & HEDGE BOOK
# ════════════════════════════════════════════
ws2 = wb.create_sheet("2-Position & Hedge")
ws2.sheet_properties.tabColor = "C00000"

ws2.cell(row=1, column=1, value="PHYSICAL POSITIONS (LONG)").font = Font(bold=True, size=11, color="2F5496")
phys_headers = ["Cargo #", "Product", "Volume (MT)", "Load Date", "Pricing Formula",
                "Pricing Month", "Contracted Price", "Current AFEI", "Unrealized P&L ($/mt)",
                "Unrealized P&L ($)", "Status"]
for c, h in enumerate(phys_headers, 1):
    ws2.cell(row=2, column=c, value=h)
style_header(ws2, 2, len(phys_headers))
for r in range(3, 20):
    ws2.cell(row=r, column=9).value = f"=IF(H{r}<>\"\",G{r}-H{r},\"\")"
    ws2.cell(row=r, column=9).fill = CALC_FILL
    ws2.cell(row=r, column=10).value = f'=IF(I{r}<>"",I{r}*C{r},"")'
    ws2.cell(row=r, column=10).fill = CALC_FILL
    ws2.cell(row=r, column=10).number_format = '#,##0'
style_range(ws2, 3, 19, len(phys_headers), INPUT_FILL)
for r in range(3, 20):
    for c in [9, 10]:
        ws2.cell(row=r, column=c).fill = CALC_FILL

ws2.cell(row=22, column=1, value="PAPER POSITIONS (SHORT — AFEI SWAPS)").font = Font(bold=True, size=11, color="2F5496")
paper_headers = ["Trade #", "Product", "Volume (MT)", "Swap Month", "Execution Price ($/mt)",
                 "Current AFEI", "Unrealized P&L ($/mt)", "Unrealized P&L ($)", "Broker", "Hedge Ref"]
for c, h in enumerate(paper_headers, 1):
    ws2.cell(row=23, column=c, value=h)
style_header(ws2, 23, len(paper_headers))
for r in range(24, 40):
    ws2.cell(row=r, column=7).value = f'=IF(F{r}<>"",E{r}-F{r},"")'
    ws2.cell(row=r, column=7).fill = CALC_FILL
    ws2.cell(row=r, column=8).value = f'=IF(G{r}<>"",G{r}*C{r},"")'
    ws2.cell(row=r, column=8).fill = CALC_FILL
    ws2.cell(row=r, column=8).number_format = '#,##0'
style_range(ws2, 24, 39, len(paper_headers), INPUT_FILL)
for r in range(24, 40):
    for c in [7, 8]:
        ws2.cell(row=r, column=c).fill = CALC_FILL

ws2.cell(row=42, column=1, value="NET EXPOSURE SUMMARY").font = Font(bold=True, size=11, color="2F5496")
net_headers = ["Month", "Physical Long (MT)", "Paper Short (MT)", "Net Open (MT)", "Hedge Ratio %", "Action Needed"]
for c, h in enumerate(net_headers, 1):
    ws2.cell(row=43, column=c, value=h)
style_header(ws2, 43, len(net_headers))
months = ["May-26", "Jun-26", "Jul-26", "Aug-26", "Sep-26", "Oct-26"]
for i, m in enumerate(months):
    r = 44 + i
    ws2.cell(row=r, column=1, value=m)
    ws2.cell(row=r, column=4).value = f"=B{r}-C{r}"
    ws2.cell(row=r, column=4).fill = CALC_FILL
    ws2.cell(row=r, column=5).value = f'=IF(B{r}>0,C{r}/B{r},"")'
    ws2.cell(row=r, column=5).fill = CALC_FILL
    ws2.cell(row=r, column=5).number_format = '0%'
    ws2.cell(row=r, column=6).value = f'=IF(AND(E{r}<>"",E{r}<0.8),"UNDER-HEDGED",IF(AND(E{r}<>"",E{r}>1.05),"OVER-HEDGED","OK"))'
    ws2.cell(row=r, column=6).fill = CALC_FILL
style_range(ws2, 44, 49, len(net_headers), INPUT_FILL)
for r in range(44, 50):
    for c in [4, 5, 6]:
        ws2.cell(row=r, column=c).fill = CALC_FILL

set_col_widths(ws2, [10, 10, 12, 14, 16, 14, 16, 16, 12, 12])
ws2.freeze_panes = "A3"


# ════════════════════════════════════════════
# 3. DAILY PRICE TRACKER
# ════════════════════════════════════════════
ws3 = wb.create_sheet("3-Price Tracker")
ws3.sheet_properties.tabColor = "00B050"

price_headers = [
    "Date",
    "Brent ($/bbl)", "Nymex CL ($/bbl)",
    "MB Propane (¢/gal)", "MB Butane (¢/gal)",
    "FEI Propane ($/mt)", "FEI Butane ($/mt)",
    "AFEI Propane ($/mt)", "AFEI Butane ($/mt)",
    "Saudi CP Propane ($/mt)", "Saudi CP Butane ($/mt)",
    "CP-AFEI Spread P", "CP-AFEI Spread B",
    "C3-C4 Spread",
    "S.China Delivered ($/mt)", "Vietnam Delivered ($/mt)",
    "Notes"
]
for c, h in enumerate(price_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(price_headers))

for r in range(2, 260):  # ~1 year of trading days
    ws3.cell(row=r, column=12).value = f'=IF(AND(J{r}<>"",H{r}<>""),J{r}-H{r},"")'  # CP-AFEI P
    ws3.cell(row=r, column=12).fill = CALC_FILL
    ws3.cell(row=r, column=13).value = f'=IF(AND(K{r}<>"",I{r}<>""),K{r}-I{r},"")'  # CP-AFEI B
    ws3.cell(row=r, column=13).fill = CALC_FILL
    ws3.cell(row=r, column=14).value = f'=IF(AND(H{r}<>"",I{r}<>""),H{r}-I{r},"")'  # C3-C4
    ws3.cell(row=r, column=14).fill = CALC_FILL

set_col_widths(ws3, [12, 13, 13, 14, 14, 14, 14, 14, 14, 15, 15, 14, 14, 12, 16, 16, 20])
ws3.freeze_panes = "B2"


# ════════════════════════════════════════════
# 4. NETBACK CALCULATOR
# ════════════════════════════════════════════
ws4 = wb.create_sheet("4-Netback Calc")
ws4.sheet_properties.tabColor = "ED7D31"

ws4.cell(row=1, column=1, value="NETBACK CALCULATOR — Pressurized Cargo to SE Asia").font = Font(bold=True, size=12, color="2F5496")

nb_headers = [
    "Scenario", "Buyer", "Destination", "Product",
    "Parcel (MT)", "CFR Offer ($/mt)",
    "PGC Freight ($/mt)", "Port Costs ($/mt)", "Insurance ($/mt)",
    "Demurrage Buffer ($/mt)", "Agent Fee ($/mt)",
    "FOB Netback ($/mt)", "Total Netback ($)",
    "vs AFEI Prem/Disc", "Rank"
]
for c, h in enumerate(nb_headers, 1):
    ws4.cell(row=2, column=c, value=h)
style_header(ws4, 2, len(nb_headers))

for r in range(3, 20):
    # FOB Netback = CFR - Freight - Port - Insurance - Demurrage - Agent
    ws4.cell(row=r, column=12).value = f"=IF(F{r}<>\"\",F{r}-G{r}-H{r}-I{r}-J{r}-K{r},\"\")"
    ws4.cell(row=r, column=12).fill = CALC_FILL
    ws4.cell(row=r, column=12).number_format = '#,##0.00'
    # Total Netback = FOB Netback * Volume
    ws4.cell(row=r, column=13).value = f'=IF(L{r}<>"",L{r}*E{r},"")'
    ws4.cell(row=r, column=13).fill = CALC_FILL
    ws4.cell(row=r, column=13).number_format = '#,##0'
    # Rank
    ws4.cell(row=r, column=15).value = f'=IF(L{r}<>"",RANK(L{r},$L$3:$L$19,0),"")'
    ws4.cell(row=r, column=15).fill = CALC_FILL

# Pre-fill reference freight ranges
ws4.cell(row=22, column=1, value="REFERENCE FREIGHT RATES (PGC)").font = Font(bold=True, size=11, color="2F5496")
ref_headers = ["Route", "Distance (days)", "Low ($/mt)", "Mid ($/mt)", "High ($/mt)", "Last Updated"]
for c, h in enumerate(ref_headers, 1):
    ws4.cell(row=23, column=c, value=h)
style_header(ws4, 23, len(ref_headers))

routes = [
    ("Brunei → South China", "2-4", 40, 50, 60),
    ("Brunei → Vietnam", "2-3", 35, 45, 55),
    ("Brunei → Philippines", "2-5", 45, 55, 70),
    ("Brunei → Bangladesh", "3-5", 50, 65, 80),
    ("Brunei → Sri Lanka", "4-6", 55, 70, 85),
    ("Singapore → Vietnam", "2-3", 30, 40, 50),
    ("Singapore → Philippines", "3-5", 40, 50, 65),
]
for i, (route, dist, lo, mid, hi) in enumerate(routes):
    r = 24 + i
    ws4.cell(row=r, column=1, value=route)
    ws4.cell(row=r, column=2, value=dist)
    ws4.cell(row=r, column=3, value=lo)
    ws4.cell(row=r, column=4, value=mid)
    ws4.cell(row=r, column=5, value=hi)
    for c in range(1, 7):
        ws4.cell(row=r, column=c).border = THIN_BORDER
        ws4.cell(row=r, column=c).font = NORMAL_FONT

style_range(ws4, 3, 19, len(nb_headers), INPUT_FILL)
for r in range(3, 20):
    for c in [12, 13, 15]:
        ws4.cell(row=r, column=c).fill = CALC_FILL

set_col_widths(ws4, [10, 16, 16, 10, 12, 14, 14, 14, 12, 16, 12, 14, 14, 16, 8])
ws4.freeze_panes = "A3"


# ════════════════════════════════════════════
# 5. CUSTOMER LOG
# ════════════════════════════════════════════
ws5 = wb.create_sheet("5-Customer Log")
ws5.sheet_properties.tabColor = "7030A0"

cust_headers = [
    "Company", "Contact Person", "Role", "Phone/IM",
    "Country", "Port(s)", "Term / Spot", "Pricing Preference",
    "Typical Volume (MT/mo)", "Credit Terms", "L/C Required?",
    "Last Deal Date", "Last Deal Price", "Last Deal Volume",
    "Relationship Score (1-5)", "Notes"
]
for c, h in enumerate(cust_headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(cust_headers))
style_range(ws5, 2, 40, len(cust_headers), INPUT_FILL)

dv_term = DataValidation(type="list", formula1='"Term,Spot,Both"', allow_blank=True)
ws5.add_data_validation(dv_term)
dv_term.add("G2:G40")

dv_lc = DataValidation(type="list", formula1='"Yes,No,Case-by-case"', allow_blank=True)
ws5.add_data_validation(dv_lc)
dv_lc.add("K2:K40")

set_col_widths(ws5, [18, 14, 12, 14, 10, 14, 10, 18, 16, 14, 12, 14, 14, 14, 16, 24])
ws5.freeze_panes = "A2"


# ════════════════════════════════════════════
# 6. P&L ATTRIBUTION
# ════════════════════════════════════════════
ws6 = wb.create_sheet("6-PnL Attribution")
ws6.sheet_properties.tabColor = "FF0000"

pnl_headers = [
    "Cargo #", "Month", "Product", "Volume (MT)",
    "Gross Revenue ($)", "Physical Cost ($)",
    "Flat Price P&L ($/mt)", "Basis P&L ($/mt)",
    "Freight P&L ($/mt)", "Timing P&L ($/mt)",
    "Total P&L ($/mt)", "Total P&L ($)",
    "Flat Price %", "Basis %", "Freight %", "Timing %",
    "Lessons / Notes"
]
for c, h in enumerate(pnl_headers, 1):
    ws6.cell(row=1, column=c, value=h)
style_header(ws6, 1, len(pnl_headers))

for r in range(2, 30):
    ws6.cell(row=r, column=11).value = f"=IF(G{r}<>\"\",G{r}+H{r}+I{r}+J{r},\"\")"
    ws6.cell(row=r, column=11).fill = CALC_FILL
    ws6.cell(row=r, column=12).value = f'=IF(K{r}<>"",K{r}*D{r},"")'
    ws6.cell(row=r, column=12).fill = CALC_FILL
    ws6.cell(row=r, column=12).number_format = '#,##0'
    # Percentage breakdown
    for pct_col, src_col in [(13, 7), (14, 8), (15, 9), (16, 10)]:
        ws6.cell(row=r, column=pct_col).value = (
            f'=IF(AND(K{r}<>"",K{r}<>0),{get_column_letter(src_col)}{r}/K{r},"")'
        )
        ws6.cell(row=r, column=pct_col).fill = CALC_FILL
        ws6.cell(row=r, column=pct_col).number_format = '0%'

style_range(ws6, 2, 29, len(pnl_headers), INPUT_FILL)
for r in range(2, 30):
    for c in [11, 12, 13, 14, 15, 16]:
        ws6.cell(row=r, column=c).fill = CALC_FILL

set_col_widths(ws6, [10, 10, 10, 12, 14, 14, 14, 14, 14, 14, 14, 14, 10, 10, 10, 10, 24])
ws6.freeze_panes = "A2"


# ════════════════════════════════════════════
# 7. FREIGHT / SHIP TRACKER
# ════════════════════════════════════════════
ws7 = wb.create_sheet("7-Freight & Ships")
ws7.sheet_properties.tabColor = "0070C0"

ws7.cell(row=1, column=1, value="PGC VESSEL AVAILABILITY").font = Font(bold=True, size=12, color="2F5496")
ship_headers = [
    "Vessel Name", "Owner/Operator", "Capacity (cbm)", "Type",
    "Current Position", "Open Date", "Last Cargo",
    "TC Rate ($/day)", "Voyage Rate ($/mt est.)", "Broker Contact", "Notes"
]
for c, h in enumerate(ship_headers, 1):
    ws7.cell(row=2, column=c, value=h)
style_header(ws7, 2, len(ship_headers))
style_range(ws7, 3, 25, len(ship_headers), INPUT_FILL)

dv_type = DataValidation(type="list", formula1='"Fully Pressurized,Semi-Pressurized,Semi-Ref"', allow_blank=True)
ws7.add_data_validation(dv_type)
dv_type.add("D3:D25")

ws7.cell(row=28, column=1, value="FREIGHT RATE LOG").font = Font(bold=True, size=12, color="2F5496")
frt_headers = ["Date", "Route", "Vessel Size (cbm)", "Rate Type", "Rate ($/mt or $/day)", "Broker", "Source"]
for c, h in enumerate(frt_headers, 1):
    ws7.cell(row=29, column=c, value=h)
style_header(ws7, 29, len(frt_headers))
style_range(ws7, 30, 60, len(frt_headers), INPUT_FILL)

dv_rate_type = DataValidation(type="list", formula1='"Voyage $/mt,TC $/day,Lumpsum"', allow_blank=True)
ws7.add_data_validation(dv_rate_type)
dv_rate_type.add("D30:D60")

set_col_widths(ws7, [16, 18, 14, 16, 16, 12, 14, 14, 16, 16, 20])
ws7.freeze_panes = "A3"


# ════════════════════════════════════════════
# 8. MARKET INTEL
# ════════════════════════════════════════════
ws8 = wb.create_sheet("8-Market Intel")
ws8.sheet_properties.tabColor = "BF8F00"

ws8.cell(row=1, column=1, value="BUYER-SIDE DEMAND SIGNALS & MARKET INTELLIGENCE").font = Font(bold=True, size=12, color="2F5496")

intel_headers = ["Date", "Market", "Category", "Signal / Event", "Source", "Impact on Our Sales", "Action Taken"]
for c, h in enumerate(intel_headers, 1):
    ws8.cell(row=2, column=c, value=h)
style_header(ws8, 2, len(intel_headers))

dv_market = DataValidation(
    type="list",
    formula1='"China PDH,China Domestic,Vietnam,Philippines,Bangladesh,Myanmar,Sri Lanka,Indonesia,India,Freight,Macro,Other"',
    allow_blank=True,
)
ws8.add_data_validation(dv_market)
dv_market.add("B3:B100")

dv_cat = DataValidation(
    type="list",
    formula1='"Demand,Supply,Price,Policy,Infrastructure,Tender,Cargo Flow,Weather,Geopolitical"',
    allow_blank=True,
)
ws8.add_data_validation(dv_cat)
dv_cat.add("C3:C100")

style_range(ws8, 3, 50, len(intel_headers), INPUT_FILL)

# Key metrics sub-table
ws8.cell(row=53, column=1, value="KEY DEMAND METRICS (update weekly)").font = Font(bold=True, size=11, color="2F5496")
metric_headers = ["Metric", "Latest Value", "Prior Value", "Change", "Date Updated", "Source"]
for c, h in enumerate(metric_headers, 1):
    ws8.cell(row=54, column=c, value=h)
style_header(ws8, 54, len(metric_headers))

metrics = [
    "China PDH Operating Rate (%)",
    "China PDH Margin ($/mt)",
    "Vietnam PV Gas Tender Calendar",
    "Philippines Petron Import Schedule",
    "Bangladesh Mongla Monthly Arrivals (MT)",
    "Indonesia Pertamina Tender",
    "AFEI M1-M2 Spread ($/mt)",
    "PGC Spot Freight ($/mt, Brunei-SCH)",
    "Brent Crude ($/bbl)",
    "USD/CNY Rate",
]
for i, met in enumerate(metrics):
    r = 55 + i
    ws8.cell(row=r, column=1, value=met)
    ws8.cell(row=r, column=4).value = f'=IF(AND(B{r}<>"",C{r}<>""),B{r}-C{r},"")'
    ws8.cell(row=r, column=4).fill = CALC_FILL
    for c in range(1, 7):
        ws8.cell(row=r, column=c).border = THIN_BORDER
        ws8.cell(row=r, column=c).font = NORMAL_FONT

set_col_widths(ws8, [12, 18, 14, 28, 18, 20, 20])
ws8.freeze_panes = "A3"


# ── Save ──
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, "LPG_Trader_Working_Sheets.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
